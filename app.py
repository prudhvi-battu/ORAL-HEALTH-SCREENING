from flask import Flask, request, jsonify, send_file, send_from_directory, render_template, redirect, session 
import tensorflow as tf
import torch
import torchvision.transforms as transforms 
import numpy as np
import cv2
from PIL import Image
import os
import openpyxl  # <-- For Excel writing
from unet_model import UNet  # Import U-Net model
from dataset import transform  # Import the same transformations from dataset.py

app = Flask(__name__, static_url_path='', static_folder='.', template_folder='.')
# Needed so we can store data in session
app.secret_key = "ORAL_HEALTH"  # choose any strong secret key
# ✅ Load Classification Model (Keras .h5)
classification_model = tf.keras.models.load_model("mobilenet_dental_model.h5")

# ✅ Load Segmentation Model (PyTorch U-Net)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
segmentation_model = UNet(in_channels=3, out_channels=5)
segmentation_model.load_state_dict(torch.load("unet_model_512.pth", map_location=device))
segmentation_model.to(device)
segmentation_model.eval()

# ✅ Create Output Directory for Segmented Images
output_dir = "output_predictions"
os.makedirs(output_dir, exist_ok=True)

# ✅ Class Labels for Classification
class_labels = ["Calculus", "Caries", "Gingivitis", "Healthy"]

# ✅ Recommendations for Each Disease
recommendations = {
    "Calculus": [
        "Professional dental cleaning is recommended.",
        "Use tartar-control toothpaste.",
        "Improve brushing and flossing habits."
    ],
    "Caries": [
        "Reduce sugar intake.",
        "Use fluoride toothpaste.",
        "Consider dental sealants."
    ],
    "Gingivitis": [
        "Brush and floss regularly.",
        "Use an antiseptic mouthwash.",
        "Schedule a professional cleaning."
    ],
    "Healthy": [
        "Maintain regular dental check-ups.",
        "Continue good oral hygiene.",
        "Eat a balanced diet for healthy teeth."
    ]
}

# ✅ Function to Map Grayscale Mask Values to Class Labels
def fix_mask_values(mask):
    mask[mask == 100] = 1  # Gingivitis → Class 1 (Red)
    mask[mask == 50] = 2   # Calculus → Class 2 (Blue)
    mask[mask == 200] = 3  # Caries → Class 3 & 4 (Green)
    return mask

# ✅ Function to Create a Colorized Mask
def create_color_mask(mask):
    mask = fix_mask_values(mask)  # Convert grayscale values to class labels
    color_mask = np.zeros((*mask.shape, 3), dtype=np.uint8)
    
    class_colors = {
        1: (255, 0, 0),   # Gingivitis → Bright Red
        2: (0, 0, 255),   # Calculus → Bright Blue
        3: (0, 255, 0),   # Caries → Bright Green
    }

    for class_id, color in class_colors.items():
        color_mask[mask == class_id] = color
    
    return color_mask

@app.route('/')
def home():
    return send_file("index.html")

@app.route('/classification.html')
def classification_page():
    return send_file("classification.html")

@app.route('/segmentation.html')
def segmentation_page():
    return send_file("segmentation.html")

@app.route('/styles.css')
def serve_css():
    return send_file("styles.css", mimetype="text/css")

# ============== Existing Classification Endpoint ==============
@app.route('/classify', methods=['POST'])
def classify():
    if 'image' not in request.files:
        return jsonify({'error': 'No image uploaded'}), 400

    file = request.files['image']
    image = Image.open(file).convert("RGB")  # Ensure 3 channels
    image = image.resize((224, 224))         # Resize to match model input
    image_array = np.array(image) / 255.0    # Normalize
    image_array = np.expand_dims(image_array, axis=0)  # Add batch dimension
    
    # ✅ Get Classification Result
    prediction = classification_model.predict(image_array)
    predicted_class = class_labels[np.argmax(prediction)]
    recommendation_list = recommendations[predicted_class]

    return jsonify({
        'disease': predicted_class,
        'recommendations': recommendation_list,
        'consult_message': "Consult a dentist for further diagnosis." if predicted_class != "Healthy" else ""
    })

# ============== Existing Segmentation Endpoint ==============
@app.route('/segment', methods=['POST'])
def segment():
    if 'image' not in request.files:
        return jsonify({'error': 'No image uploaded'}), 400

    file = request.files['image']
    image = Image.open(file).convert("RGB")
    image_np = np.array(image)
    
    # ✅ Resize and apply same preprocessing
    transformed = transform(image=image_np)
    image_tensor = transformed["image"].unsqueeze(0).to(device)

    # ✅ Get Segmentation Mask
    with torch.no_grad():
        output = segmentation_model(image_tensor)
        predicted_mask = torch.argmax(output, dim=1).squeeze().cpu().numpy()

    # ✅ Resize mask back to original
    mask_resized = cv2.resize(predicted_mask, (image_np.shape[1], image_np.shape[0]), interpolation=cv2.INTER_NEAREST)

    # ✅ Create Colorized Mask
    color_mask = create_color_mask(mask_resized)

    # ✅ Save Input & Segmentation Mask
    input_filename = os.path.join(output_dir, file.filename)
    output_filename = os.path.join(output_dir, file.filename.replace(".", "_segmented."))

    image.save(input_filename)
    cv2.imwrite(output_filename, cv2.cvtColor(color_mask, cv2.COLOR_RGB2BGR))

    return jsonify({
        'input_image': f"/output_predictions/{os.path.basename(input_filename)}",
        'segmentation_result': f"/output_predictions/{os.path.basename(output_filename)}"
    })

@app.route('/output_predictions/<filename>')
def get_segmented_image(filename):
    return send_from_directory(output_dir, filename)

# ============== NEW: Store Help Form Submissions in help_records.xlsx ==============
@app.route('/submit_help_form', methods=['POST'])
def submit_help_form():
    # Extract form data
    name = request.form.get("name")
    email = request.form.get("email")
    phone = request.form.get("phone", "")
    message = request.form.get("message")

    # Path to the Excel file (for help.html)
    excel_file = "help_records.xlsx"

    # Check if file exists; if not, create it with headers
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HelpRequests"
        ws.append(["Name", "Email", "Phone", "Message"])
        wb.save(excel_file)

    # Open and append new row
    wb = openpyxl.load_workbook(excel_file)
    ws = wb["HelpRequests"]
    ws.append([name, email, phone, message])
    wb.save(excel_file)

    # Redirect or show a confirmation page
    return "Thank you for contacting us! <br><a href='help.html'>Go back</a>"

# ============== NEW: Store ind2.html Form Submissions in ind2_records.xlsx ==============
@app.route('/submit_ind2_form', methods=['POST'])
def submit_ind2_form():
    user_name = request.form.get("userName")
    user_age = request.form.get("userAge")
    user_gender = request.form.get("userGender")
    symptoms = request.form.getlist("symptoms")
    others_text = request.form.get("othersText", "").strip()

    # If user checked "Others" and typed in custom text, incorporate that
    if "Others" in symptoms:
        symptoms.remove("Others")
        if others_text:
            symptoms.append(others_text)

    symptoms_joined = ", ".join(symptoms) if symptoms else ""

    excel_file = "ind2_records.xlsx"
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ind2Submissions"
        # Add an extra column for the user’s choice
        ws.append(["Name", "Age", "Gender", "Symptoms", "OptionClicked"])
        wb.save(excel_file)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb["Ind2Submissions"]

    # Append the user’s data: leave OptionClicked blank for now
    ws.append([user_name, user_age, user_gender, symptoms_joined, ""])
    
    # Grab the row number we just wrote to
    new_row = ws.max_row  # the last row in the sheet

    wb.save(excel_file)

    # Store that row in session so we can update it later
    session["last_ind2_row"] = new_row

    # Now redirect user to option.html
    return redirect("option.html")

@app.route("/record_option")
def record_option():
    chosen_option = request.args.get("option", "")  # "classification" or "segmentation"
    excel_file = "ind2_records.xlsx"

    row_num = session.get("last_ind2_row", None)
    if row_num is not None and os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Ind2Submissions"]

        # row_num is an integer; update the cell in the "OptionClicked" column
        # We made OptionClicked the 5th column in our sheet, so column=5
        ws.cell(row=row_num, column=5).value = chosen_option
        wb.save(excel_file)

    # Now decide where to redirect them
    if chosen_option == "classification":
        return redirect("/classification.html")
    elif chosen_option == "segmentation":
        return redirect("/segmentation.html")
    else:
        # If something else, go back to option or home
        return redirect("/option.html")

if __name__ == '__main__':
    app.run(debug=True)